from sqlalchemy.orm import Session
from models import MeeshoSale, MeeshoReturn, MeeshoInvoice
from collections import defaultdict
import csv
import pandas as pd
import os
from datetime import datetime, timedelta
from constants import (
    B2CL_INVOICE_THRESHOLD, TransactionType,
    STATE_CODE_MAPPING,
    get_state_code, generate_note_number,
    NoteType
)


def normalize_rate(rate_value):
    """
    Normalize GST rate to percentage format (whole number).
    Converts 0.05 -> 5.0, 0.12 -> 12.0, 0.18 -> 18.0, 5 -> 5.0, etc.
    """
    if rate_value is None or rate_value == 0:
        return 0.0
    
    try:
        rate = float(rate_value)
    except (ValueError, TypeError):
        return 0.0
    
    # If rate is between 0 and 1 (decimal format like 0.05, 0.12, 0.18), multiply by 100
    if 0 < rate < 1:
        rate = rate * 100
    
    # Return as 2-decimal float (e.g., 5.0, 12.0, 18.0)
    return round(rate, 2)


def get_flipkart_gst_excel_path(config_path="config.json"):
    """
    Get the Flipkart GST Excel file path from configuration.
    Users should upload/import the file and configure the path.
    Returns the file path if configured, else None.
    """
    import json
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
            return config.get('flipkart_gst_excel_path')
    except:
        return None

def set_flipkart_gst_excel_path(file_path, config_path="config.json"):
    """
    Set the Flipkart GST Excel file path in configuration.
    Call this after user uploads/selects the file through the UI.
    """
    import json
    try:
        config = {}
        try:
            with open(config_path, 'r') as f:
                config = json.load(f)
        except:
            pass
        
        config['flipkart_gst_excel_path'] = file_path
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=2)
        return True
    except Exception as e:
        print(f"Error setting Flipkart GST Excel path: {e}")
        return False

def read_flipkart_gst_b2cs_data(excel_file_path):
    """
    Read B2CS (B2C Sales) data from Flipkart's GSTR-1 Excel file.
    Uses Section 7(B)(2) which contains state-wise B2C sales with IGST.
    Returns a dictionary keyed by (state, rate) with taxable values.
    """
    try:
        df = pd.read_excel(excel_file_path, sheet_name="Section 7(B)(2) in GSTR-1")
        
        data = {}
        for _, row in df.iterrows():
            state = str(row.get('Delivered State (PoS)', '')).strip()
            rate = row.get('IGST %', 0)
            taxable_value = row.get('Aggregate Taxable Value Rs.', 0)
            
            if state and taxable_value:
                # Normalize state to match our STATE_CODE_MAPPING format
                state_upper = state.upper()
                normalized_state = STATE_CODE_MAPPING.get(state_upper, state)
                rate_normalized = normalize_rate(rate)
                
                key = (normalized_state, rate_normalized)
                data[key] = data.get(key, 0) + taxable_value
        
        return data
    except Exception as e:
        print(f"Error reading Flipkart GST B2CS data: {e}")
        return None

def read_flipkart_gst_hsn_data(excel_file_path):
    """
    Read HSN-wise summary data from Flipkart's GSTR-1 Excel file.
    Uses Section 12 which contains HSN code aggregates.
    Returns a dictionary keyed by (hsn, rate) with detailed tax information.
    """
    try:
        df = pd.read_excel(excel_file_path, sheet_name="Section 12 in GSTR-1")
        
        data = {}
        for _, row in df.iterrows():
            hsn = str(row.get('HSN Number', '')).strip()
            qty = row.get('Total Quantity in Nos.', 0)
            total_value = row.get('Total\n Value Rs.', 0)
            taxable_value = row.get('Total Taxable Value Rs.', 0)
            igst = row.get('IGST Amount Rs.', 0)
            cgst = row.get('CGST Amount Rs.', 0)
            sgst = row.get('SGST Amount Rs.', 0)
            cess = row.get('Cess Rs.', 0)
            
            if hsn and taxable_value:
                # Calculate rate from tax amounts, then snap to nearest GST slab
                if igst and igst > 0:
                    rate = (igst / taxable_value) * 100 if taxable_value else 0
                elif cgst and sgst:
                    rate = ((cgst + sgst) / taxable_value) * 100 if taxable_value else 0
                else:
                    rate = 0

                rate_normalized = approximate_gst_rate(normalize_rate(rate))
                key = (hsn, rate_normalized)
                
                data[key] = {
                    'quantity': int(qty) if qty else 0,
                    'total_value': total_value,
                    'taxable_value': taxable_value,
                    'igst_amount': igst,
                    'cgst_amount': cgst,
                    'sgst_amount': sgst,
                    'cess_amount': cess
                }
        
        return data
    except Exception as e:
        print(f"Error reading Flipkart GST HSN data: {e}")
        return None
def _extract_flipkart_excel_gstin(excel_file_path):
    """Extract the seller GSTIN from a Flipkart GST Excel file for validation."""
    try:
        xl = pd.ExcelFile(excel_file_path)
        for sheet in xl.sheet_names:
            if sheet != 'Help':
                df = pd.read_excel(excel_file_path, sheet_name=sheet, nrows=1)
                if 'GSTIN' in df.columns and not df.empty:
                    gstin_value = df['GSTIN'].iloc[0]
                    if pd.notna(gstin_value) and str(gstin_value).strip():
                        return str(gstin_value).strip()
    except Exception:
        pass
    return None

GST_SLABS = [5.0, 12.0, 18.0]

def approximate_gst_rate(raw_rate):
    return min(GST_SLABS, key=lambda x: abs(x - raw_rate))

def get_gstin_for_supplier(supplier_id: int, db: Session):
    from models import SellerMapping
    gstin_obj = db.query(MeeshoSale.gstin).filter(
        MeeshoSale.supplier_id == supplier_id,
        MeeshoSale.gstin.isnot(None),
        MeeshoSale.gstin != ""
    ).first()
    if gstin_obj and gstin_obj[0]:
        return gstin_obj[0]
    mapping = db.query(SellerMapping).filter(
        SellerMapping.supplier_id == supplier_id
    ).first()
    if mapping and mapping.gstin:
        return mapping.gstin
    return None

def _get_gst_pivot_data(financial_year: int, month_number: int, supplier_id: int, db: Session):
    sales = db.query(MeeshoSale).filter_by(
        financial_year=financial_year,
        month_number=month_number,
        supplier_id=supplier_id
    ).all()
    returns = db.query(MeeshoReturn).filter_by(
        financial_year=financial_year,
        month_number=month_number,
        supplier_id=supplier_id
    ).all()

    pivot = defaultdict(float)

    def normalize_state(state_raw: str) -> str:
        state_upper = str(state_raw or "").strip().upper()
        return STATE_CODE_MAPPING.get(state_upper, state_upper)

    for record in sales:
        pivot[(normalize_state(record.end_customer_state_new), float(record.gst_rate or 0))] += record.total_taxable_sale_value or 0

    for record in returns:
        pivot[(normalize_state(record.end_customer_state_new), float(record.gst_rate or 0))] -= record.total_taxable_sale_value or 0

    rows = [
        {"state": state, "gst_rate": gst_rate, "total_taxable_value": round(value, 2)}
        for (state, gst_rate), value in pivot.items() if state
    ]
    rows.sort(key=lambda r: (r["state"], r["gst_rate"]))
    return rows

def generate_gst_pivot_csv(financial_year, month_number, gstin_or_supplier_id, db,
                           file_path=None, output_folder=None):
    """
    Dynamic-path GST B2CS pivot generator - reads all marketplace data from database.
    
    Args:
        gstin_or_supplier_id: Either GSTIN string or legacy supplier_id integer
    """
    from models import FlipkartOrder, FlipkartReturn, AmazonOrder, AmazonReturn
    
    # Get GSTIN - accept either GSTIN string or legacy supplier_id
    if isinstance(gstin_or_supplier_id, str) and len(gstin_or_supplier_id) == 15:
        gstin = gstin_or_supplier_id
        supplier_id = None  # Not needed for new queries
    else:
        gstin = get_gstin_for_supplier(gstin_or_supplier_id, db)
        supplier_id = gstin_or_supplier_id
        if not gstin:
            raise ValueError(f"GSTIN not found for supplier ID {gstin_or_supplier_id}")

    combined_data = {}
    
    # 1. Meesho DB - query by GSTIN or supplier_id
    if isinstance(gstin_or_supplier_id, str) and len(gstin_or_supplier_id) == 15:
        # Query Meesho by GSTIN
        sales = db.query(MeeshoSale).filter_by(
            financial_year=financial_year,
            month_number=month_number,
            gstin=gstin_or_supplier_id
        ).all()
        returns = db.query(MeeshoReturn).filter_by(
            financial_year=financial_year,
            month_number=month_number,
            gstin=gstin_or_supplier_id
        ).all()
        
        # Process Meesho data
        def normalize_state(state_raw: str) -> str:
            state_upper = str(state_raw or "").strip().upper()
            return STATE_CODE_MAPPING.get(state_upper, state_upper)
        
        for record in sales:
            key = (normalize_state(record.end_customer_state_new), round(float(record.gst_rate or 0), 2))
            combined_data[key] = combined_data.get(key, 0) + (record.total_taxable_sale_value or 0)
        
        for record in returns:
            key = (normalize_state(record.end_customer_state_new), round(float(record.gst_rate or 0), 2))
            combined_data[key] = combined_data.get(key, 0) - (record.total_taxable_sale_value or 0)
    
    elif supplier_id:
        # Legacy path - use old function
        b2cs_rows = _get_gst_pivot_data(financial_year, month_number, supplier_id, db)
        for row in b2cs_rows:
            combined_data[(row["state"], round(row["gst_rate"], 2))] = combined_data.get((row["state"], round(row["gst_rate"], 2)), 0) + row["total_taxable_value"]

    # Calculate date range for Flipkart and Amazon queries
    if month_number <= 3:
        year = financial_year
    else:
        year = financial_year - 1
    month_start = datetime(year, month_number, 1)
    if month_number == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month_number + 1, 1)

    # 2. Flipkart - Use certified GST Excel data when configured (validate GSTIN match)
    flipkart_excel = get_flipkart_gst_excel_path()
    use_flipkart_excel = False
    if flipkart_excel and os.path.exists(flipkart_excel):
        excel_gstin = _extract_flipkart_excel_gstin(flipkart_excel)
        if excel_gstin and excel_gstin == gstin:
            flipkart_data = read_flipkart_gst_b2cs_data(flipkart_excel)
            if flipkart_data:
                for key, taxable_value in flipkart_data.items():
                    combined_data[key] = combined_data.get(key, 0) + taxable_value
                use_flipkart_excel = True

    if not use_flipkart_excel:
        # Use database (GSTIN-filtered)
        flipkart_query = db.query(FlipkartOrder).filter(
            FlipkartOrder.event_type == 'Sale',
            FlipkartOrder.order_date >= month_start,
            FlipkartOrder.order_date < month_end,
            FlipkartOrder.seller_gstin == gstin
        )
        flipkart_orders = flipkart_query.all()
        
        for order in flipkart_orders:
            delivery_state = str(order.customer_delivery_state or "").strip().upper()
            delivery_state_normalized = get_state_code(delivery_state)
            
            if order.igst_rate and order.igst_rate > 0:
                gst_rate = normalize_rate(order.igst_rate)
            elif order.cgst_rate and order.sgst_rate:
                gst_rate = normalize_rate(order.cgst_rate + order.sgst_rate)
            else:
                continue
            
            taxable_value = order.taxable_value or 0
            if taxable_value > 0:
                key = (delivery_state_normalized, gst_rate)
                combined_data[key] = combined_data.get(key, 0) + taxable_value

        flipkart_returns = db.query(FlipkartReturn).filter(
            FlipkartReturn.order_date >= month_start,
            FlipkartReturn.order_date < month_end,
            FlipkartReturn.seller_gstin == gstin
        ).all()
        
        for ret in flipkart_returns:
            delivery_state = str(ret.customer_delivery_state or "").strip().upper()
            delivery_state_normalized = get_state_code(delivery_state)

            if ret.igst_rate and ret.igst_rate > 0:
                gst_rate = normalize_rate(ret.igst_rate)
            elif ret.cgst_rate and ret.sgst_rate:
                gst_rate = normalize_rate(ret.cgst_rate + ret.sgst_rate)
            else:
                continue

            taxable_value = abs(float(ret.taxable_value or 0))
            if taxable_value > 0:
                key = (delivery_state_normalized, gst_rate)
                combined_data[key] = combined_data.get(key, 0) - taxable_value

    # 3. Amazon DB - aggregate by state and GST rate - B2C only (exclude B2B which goes to Table 4)
    amazon_query = db.query(AmazonOrder).filter(
        AmazonOrder.transaction_type == TransactionType.SHIPMENT,
        AmazonOrder.order_date >= month_start,
        AmazonOrder.order_date < month_end,
        AmazonOrder.seller_gstin == gstin,
        (AmazonOrder.customer_bill_to_gstid.is_(None) |
         (AmazonOrder.customer_bill_to_gstid == '') |
         (AmazonOrder.customer_bill_to_gstid == 'nan'))
    )
    amazon_orders = amazon_query.all()
    
    for order in amazon_orders:
        # Normalize ship-to state
        delivery_state_normalized = get_state_code(order.ship_to_state) if order.ship_to_state else "Unknown"
        
        # Calculate GST rate: Interstate (IGST) or Intrastate (CGST + SGST) - normalize rate
        gst_rate = None
        if order.igst_rate:
            # Interstate - use IGST rate
            gst_rate = normalize_rate(order.igst_rate)
        elif order.cgst_rate and order.sgst_rate:
            # Intrastate - use CGST + SGST
            gst_rate = normalize_rate(order.cgst_rate + order.sgst_rate)
        else:
            continue  # Skip if no tax rate
        
        taxable_value = order.taxable_value or 0
        if taxable_value > 0:
            key = (delivery_state_normalized, gst_rate)
            combined_data[key] = combined_data.get(key, 0) + taxable_value

    # Amazon returns - B2C only (exclude B2B which goes to CDNR Table 9B)
    amazon_returns = db.query(AmazonReturn).filter(
        AmazonReturn.transaction_type == TransactionType.REFUND,
        AmazonReturn.order_date >= month_start,
        AmazonReturn.order_date < month_end,
        AmazonReturn.seller_gstin == gstin,
        (AmazonReturn.customer_bill_to_gstid.is_(None) |
         (AmazonReturn.customer_bill_to_gstid == '') |
         (AmazonReturn.customer_bill_to_gstid == 'nan'))
    ).all()
    
    for ret in amazon_returns:
        # Normalize ship-to state
        delivery_state_normalized = get_state_code(ret.ship_to_state) if ret.ship_to_state else "Unknown"

        # Calculate GST rate: Interstate (IGST) or Intrastate (CGST + SGST) - use rate fields only (no approximation)
        gst_rate = None
        if ret.igst_rate:
            gst_rate = normalize_rate(ret.igst_rate)
        elif ret.cgst_rate and ret.sgst_rate:
            gst_rate = normalize_rate(ret.cgst_rate + ret.sgst_rate)
        else:
            continue  # Skip if no tax rate can be determined

        taxable_value = abs(float(ret.taxable_value or 0))
        if taxable_value > 0:
            key = (delivery_state_normalized, gst_rate)
            combined_data[key] = combined_data.get(key, 0) - taxable_value  # Subtract returns

    # 4. Output
    if not file_path:
        file_path = os.path.join(output_folder or "", "b2cs.csv")
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Type", "Place Of Supply", "Rate", "Applicable % of Tax Rate", "Taxable Value", "Cess Amount", "E-Commerce GSTIN"])
        writer.writeheader()
        for (state, gst_rate), taxable_value in sorted(combined_data.items()):
            writer.writerow({"Type": "OE", "Place Of Supply": state, "Rate": gst_rate, "Applicable % of Tax Rate": "", "Taxable Value": round(taxable_value, 2), "Cess Amount": "", "E-Commerce GSTIN": ""})
    return f"✅ Combined GST CSV written to {file_path} with {len(combined_data)} aggregated rows (Meesho + Flipkart + Amazon)."

def generate_gst_hsn_pivot_csv(financial_year, month_number, gstin_or_supplier_id, db,
                               file_path=None, output_folder=None):
    """
    Dynamic-path GST HSN pivot generator - reads all marketplace data from database.
    
    Args:
        gstin_or_supplier_id: Either GSTIN string or legacy supplier_id integer
    """
    from models import FlipkartOrder, FlipkartReturn, AmazonOrder, AmazonReturn
    
    # Get GSTIN and handle both new GSTIN and legacy supplier_id
    if isinstance(gstin_or_supplier_id, str) and len(gstin_or_supplier_id) == 15:
        gstin_for_supplier = gstin_or_supplier_id
        # Query Meesho data by GSTIN
        sales = db.query(MeeshoSale).filter_by(
            financial_year=financial_year, 
            month_number=month_number, 
            gstin=gstin_or_supplier_id
        ).all()
        returns = db.query(MeeshoReturn).filter_by(
            financial_year=financial_year, 
            month_number=month_number, 
            gstin=gstin_or_supplier_id
        ).all()
    else:
        # Legacy supplier_id path
        gstin_for_supplier = get_gstin_for_supplier(gstin_or_supplier_id, db)
        sales = db.query(MeeshoSale).filter_by(
            financial_year=financial_year, 
            month_number=month_number, 
            supplier_id=gstin_or_supplier_id
        ).all()
        returns = db.query(MeeshoReturn).filter_by(
            financial_year=financial_year, 
            month_number=month_number, 
            supplier_id=gstin_or_supplier_id
        ).all()

    # derive supplier state from supplier GSTIN (first two digits)
    supplier_state_display = None
    if gstin_for_supplier and len(gstin_for_supplier) >= 2 and gstin_for_supplier[:2].isdigit():
        code = gstin_for_supplier[:2]
        # STATE_CODE_MAPPING values are like "23-Madhya Pradesh" — find the value that starts with the code
        for v in STATE_CODE_MAPPING.values():
            try:
                if str(v).startswith(f"{int(code):02d}-"):
                    supplier_state_display = v
                    break
            except Exception:
                continue

    def normalize_state(state_raw: str) -> str:
        state_upper = str(state_raw or "").strip().upper()
        return STATE_CODE_MAPPING.get(state_upper, state_upper)

    def is_intra(state):
        """Return True if the given destination state is intrastate (same as supplier state).

        We compare normalized state representations (e.g. "23-Madhya Pradesh") when possible.
        If supplier state can't be determined from GSTIN, conservatively return False.
        """
        norm = normalize_state(state)
        if supplier_state_display:
            return norm == supplier_state_display
        return False

    pivot_data = defaultdict(lambda: {"taxable_value": 0.0, "quantity": 0, "cgst_amount": 0.0, "sgst_amount": 0.0, "igst_amount": 0.0, "cess_amount": 0.0})

    for rec in sales:
        k = (str(rec.hsn_code or "UNKNOWN"), float(rec.gst_rate or 0))
        pivot_data[k]["quantity"] += int(rec.quantity or 0)
        pivot_data[k]["taxable_value"] += float(rec.total_taxable_sale_value or 0.0)
        if is_intra(rec.end_customer_state_new):
            pivot_data[k]["cgst_amount"] += (rec.total_taxable_sale_value or 0) * rec.gst_rate / 200
            pivot_data[k]["sgst_amount"] += (rec.total_taxable_sale_value or 0) * rec.gst_rate / 200
        else:
            pivot_data[k]["igst_amount"] += (rec.total_taxable_sale_value or 0) * rec.gst_rate / 100

    for rec in returns:
        k = (str(rec.hsn_code or "UNKNOWN"), float(rec.gst_rate or 0))
        pivot_data[k]["quantity"] -= int(rec.quantity or 0)
        pivot_data[k]["taxable_value"] -= float(rec.total_taxable_sale_value or 0.0)
        if is_intra(rec.end_customer_state_new):
            pivot_data[k]["cgst_amount"] -= (rec.total_taxable_sale_value or 0) * rec.gst_rate / 200
            pivot_data[k]["sgst_amount"] -= (rec.total_taxable_sale_value or 0) * rec.gst_rate / 200
        else:
            pivot_data[k]["igst_amount"] -= (rec.total_taxable_sale_value or 0) * rec.gst_rate / 100

    # Flipkart HSN merge - now from database
    if month_number <= 3:  # Jan-Mar
        year = financial_year
    else:  # Apr-Dec
        year = financial_year - 1
    
    month_start = datetime(year, month_number, 1)
    if month_number == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month_number + 1, 1)
    
    flipkart_orders = []
    flipkart_returns = []
    hsn_rate_map = {}
    
    # Try to use Flipkart certified GST Excel data first (validate GSTIN match)
    flipkart_excel = get_flipkart_gst_excel_path()
    use_flipkart_excel = False
    if flipkart_excel and os.path.exists(flipkart_excel):
        excel_gstin = _extract_flipkart_excel_gstin(flipkart_excel)
        if excel_gstin and excel_gstin == gstin_for_supplier:
            flipkart_hsn_data = read_flipkart_gst_hsn_data(flipkart_excel)
            if flipkart_hsn_data:
                for (hsn, rate), vals in flipkart_hsn_data.items():
                    hsn_rate_map[hsn] = rate
                    k = (hsn, rate)
                    pivot_data[k]["quantity"] += vals.get('quantity', 0)
                    pivot_data[k]["taxable_value"] += vals.get('taxable_value', 0)
                    pivot_data[k]["igst_amount"] += vals.get('igst_amount', 0)
                    pivot_data[k]["cgst_amount"] += vals.get('cgst_amount', 0)
                    pivot_data[k]["sgst_amount"] += vals.get('sgst_amount', 0)
                    pivot_data[k]["cess_amount"] += vals.get('cess_amount', 0)
                use_flipkart_excel = True

    if not use_flipkart_excel:
        # Use database (GSTIN-filtered)
        flipkart_orders = db.query(FlipkartOrder).filter(
            FlipkartOrder.event_type == 'Sale',
            FlipkartOrder.order_date >= month_start,
            FlipkartOrder.order_date < month_end,
            FlipkartOrder.seller_gstin == gstin_for_supplier
        ).all()
        
        flipkart_returns = db.query(FlipkartReturn).filter(
            FlipkartReturn.order_date >= month_start,
            FlipkartReturn.order_date < month_end,
            FlipkartReturn.seller_gstin == gstin_for_supplier
        ).all()
        
        # Aggregate Flipkart sales by HSN
        for order in flipkart_orders:
            hsn = str(order.hsn_code or "UNKNOWN")
            if order.igst_rate and order.igst_rate > 0:
                rate = normalize_rate(order.igst_rate)
            elif order.cgst_rate and order.sgst_rate:
                rate = normalize_rate(order.cgst_rate + order.sgst_rate)
            else:
                rate = 0
            
            if hsn not in hsn_rate_map:
                hsn_rate_map[hsn] = rate
            
            k = (hsn, rate)
            pivot_data[k]["quantity"] += int(order.quantity or 0)
            pivot_data[k]["taxable_value"] += float(order.taxable_value or 0)
            pivot_data[k]["igst_amount"] += float(order.igst_amount or 0)
            pivot_data[k]["cgst_amount"] += float(order.cgst_amount or 0)
            pivot_data[k]["sgst_amount"] += float(order.sgst_amount or 0)
        
        # Subtract Flipkart returns (use abs() since returns may store negative amounts)
        for ret in flipkart_returns:
            hsn = str(ret.hsn_code or "UNKNOWN")

            if hsn in hsn_rate_map:
                rate = hsn_rate_map[hsn]
            else:
                if ret.igst_rate and ret.igst_rate > 0:
                    rate = normalize_rate(ret.igst_rate)
                elif ret.cgst_rate and ret.sgst_rate:
                    rate = normalize_rate(ret.cgst_rate + ret.sgst_rate)
                else:
                    rate = 0.0
                if hsn not in hsn_rate_map:
                    hsn_rate_map[hsn] = rate

            k = (hsn, rate)
            pivot_data[k]["quantity"] -= abs(int(ret.quantity or 0))
            pivot_data[k]["taxable_value"] -= abs(float(ret.taxable_value or 0))
            pivot_data[k]["igst_amount"] -= abs(float(ret.igst_amount or 0))
            pivot_data[k]["cgst_amount"] -= abs(float(ret.cgst_amount or 0))
            pivot_data[k]["sgst_amount"] -= abs(float(ret.sgst_amount or 0))

    # Amazon HSN merge - B2C only (exclude B2B which goes to HSN B2B report)
    amazon_orders = db.query(AmazonOrder).filter(
        AmazonOrder.transaction_type == TransactionType.SHIPMENT,
        AmazonOrder.order_date >= month_start,
        AmazonOrder.order_date < month_end,
        AmazonOrder.seller_gstin == gstin_for_supplier,
        (AmazonOrder.customer_bill_to_gstid.is_(None) |
         (AmazonOrder.customer_bill_to_gstid == '') |
         (AmazonOrder.customer_bill_to_gstid == 'nan'))
    ).all()

    amazon_returns = db.query(AmazonReturn).filter(
        AmazonReturn.transaction_type == TransactionType.REFUND,
        AmazonReturn.order_date >= month_start,
        AmazonReturn.order_date < month_end,
        AmazonReturn.seller_gstin == gstin_for_supplier,
        (AmazonReturn.customer_bill_to_gstid.is_(None) |
         (AmazonReturn.customer_bill_to_gstid == '') |
         (AmazonReturn.customer_bill_to_gstid == 'nan'))
    ).all()
    
    # Aggregate Amazon sales by HSN (update the map)
    for order in amazon_orders:
        hsn = str(order.hsn_sac or "UNKNOWN")
        # Calculate effective tax rate and normalize
        if order.igst_rate and order.igst_rate > 0:
            rate = normalize_rate(order.igst_rate)
        elif order.cgst_rate and order.sgst_rate:
            rate = normalize_rate(order.cgst_rate + order.sgst_rate)
        else:
            rate = 0
        
        # Store the rate for this HSN if not already set
        if hsn not in hsn_rate_map:
            hsn_rate_map[hsn] = rate
        
        k = (hsn, rate)
        pivot_data[k]["quantity"] += int(order.quantity or 0)
        pivot_data[k]["taxable_value"] += float(order.taxable_value or 0)
        pivot_data[k]["igst_amount"] += float(order.igst_amount or 0)
        pivot_data[k]["cgst_amount"] += float(order.cgst_amount or 0)
        pivot_data[k]["sgst_amount"] += float(order.sgst_amount or 0)
    
    # Subtract Amazon returns (use abs() since returns may store negative amounts)
    for ret in amazon_returns:
        hsn = str(ret.hsn_sac or "UNKNOWN")

        # Use the rate established from sales (most reliable source)
        if hsn in hsn_rate_map:
            rate = hsn_rate_map[hsn]
        else:
            # For returns without prior sales, use the return's own rate fields (NORMALIZED)
            if ret.igst_rate and ret.igst_rate > 0:
                rate = normalize_rate(ret.igst_rate)
            elif ret.cgst_rate and ret.sgst_rate:
                rate = normalize_rate(ret.cgst_rate + ret.sgst_rate)
            else:
                rate = 0.0
            # Store this rate for future use
            if hsn not in hsn_rate_map:
                hsn_rate_map[hsn] = rate

        k = (hsn, rate)
        pivot_data[k]["quantity"] -= abs(int(ret.quantity or 0))
        pivot_data[k]["taxable_value"] -= abs(float(ret.taxable_value or 0))
        pivot_data[k]["igst_amount"] -= abs(float(ret.igst_amount or 0))
        pivot_data[k]["cgst_amount"] -= abs(float(ret.cgst_amount or 0))
        pivot_data[k]["sgst_amount"] -= abs(float(ret.sgst_amount or 0))

    # Output - use pivot_data directly without consolidation
    # The hsn_rate_map ensures each HSN primarily uses one rate from sales
    if not file_path:
        file_path = os.path.join(output_folder or "", "hsn(b2c).csv")
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["HSN", "Description", "UQC", "Total Quantity", "Total Value", "Taxable Value", "Integrated Tax Amount", "Central Tax Amount", "State/UT Tax Amount", "Cess Amount", "Rate"])
        for (hsn, rate), vals in sorted(pivot_data.items()):
            # Skip rows with zero quantity (no transactions)
            if vals["quantity"] == 0:
                continue
            
            total_value = round(vals["taxable_value"] + vals["igst_amount"] + vals["cgst_amount"] + vals["sgst_amount"] + vals["cess_amount"], 2)
            # Ensure rate is normalized before output
            normalized_rate = normalize_rate(rate) if rate else 0
            writer.writerow([hsn, "", "NOS-NUMBERS", vals["quantity"], total_value, round(vals["taxable_value"], 2), round(vals["igst_amount"], 2), round(vals["cgst_amount"], 2), round(vals["sgst_amount"], 2), round(vals["cess_amount"], 2), normalized_rate])
    return f"✅ GST HSN pivot CSV saved as '{file_path}' for FY {financial_year}, Month {month_number}, GSTIN {gstin_or_supplier_id}"


def generate_b2b_csv(financial_year, month_number, gstin_or_supplier_id, db,
                     file_path=None, output_folder=None):
    """
    Generate B2B invoice-level CSV from Amazon B2B transactions (where customer_bill_to_gstid is present).
    Complies with GSTR-1 Table 4A, 4B, 4C format with rate-wise breakdown.
    
    Args:
        gstin_or_supplier_id: Either GSTIN string or legacy supplier_id integer
    
    GSTR-1 Compliance:
    - Invoice-level details for all B2B supplies (Table 4A)
    - Rate-wise breakdown (multiple rows per invoice if different GST rates)
    - Includes recipient GSTIN for ITC claim
    - Place of Supply for destination-based taxation
    
    Format: GSTIN of Supplier, Trade/Legal name, Receiver GSTIN, Invoice Number, Invoice date,
            Invoice Value, Place of Supply, Reverse Charge, Invoice Type, E-Commerce GSTIN, 
            Rate, Taxable Value, Cess Amount
    """
    from models import AmazonOrder, AmazonReturn
    from datetime import datetime
    
    # Calculate month start and end dates
    if month_number <= 3:  # Jan-Mar
        year = financial_year
    else:  # Apr-Dec
        year = financial_year - 1
    
    month_start = datetime(year, month_number, 1)
    if month_number == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month_number + 1, 1)

    # Resolve supplier GSTIN early for filtering
    if isinstance(gstin_or_supplier_id, str) and len(gstin_or_supplier_id) == 15:
        supplier_gstin = gstin_or_supplier_id
    else:
        supplier_gstin = get_gstin_for_supplier(gstin_or_supplier_id, db)
        if not supplier_gstin:
            raise ValueError(f"GSTIN not found for supplier ID {gstin_or_supplier_id}")

    # Query B2B transactions (where customer GSTIN is present) - filtered by seller GSTIN
    b2b_orders = db.query(AmazonOrder).filter(
        AmazonOrder.transaction_type == TransactionType.SHIPMENT,
        AmazonOrder.customer_bill_to_gstid.isnot(None),
        AmazonOrder.customer_bill_to_gstid != '',
        AmazonOrder.customer_bill_to_gstid != 'nan',
        AmazonOrder.order_date >= month_start,
        AmazonOrder.order_date < month_end,
        AmazonOrder.seller_gstin == supplier_gstin
    ).all()

    b2b_returns = db.query(AmazonReturn).filter(
        AmazonReturn.customer_bill_to_gstid.isnot(None),
        AmazonReturn.customer_bill_to_gstid != '',
        AmazonReturn.customer_bill_to_gstid != 'nan',
        AmazonReturn.order_date >= month_start,
        AmazonReturn.order_date < month_end,
        AmazonReturn.seller_gstin == supplier_gstin
    ).all()

    # Group by (invoice_no, rate) to create rate-wise breakdown per invoice (GSTR-1 requirement)
    invoice_data = {}

    for order in b2b_orders:
        invoice_no = order.invoice_number or "UNKNOWN"
        
        # Calculate GST rate for this order
        if order.igst_rate and order.igst_rate > 0:
            rate = round(order.igst_rate, 2)
        elif order.cgst_rate and order.sgst_rate:
            rate = round(order.cgst_rate + order.sgst_rate, 2)
        else:
            rate = 0
        
        # Key: (invoice_no, rate) for rate-wise breakdown
        key = (invoice_no, rate)
        
        if key not in invoice_data:
            invoice_data[key] = {
                'receiver_gstin': order.customer_bill_to_gstid or "",
                'receiver_name': order.buyer_name or "",
                'invoice_date': order.invoice_date,
                'place_of_supply': get_state_code(order.bill_to_state or order.ship_to_state),
                'invoice_value': 0,
                'taxable_value': 0,
                'igst_amount': 0,
                'cgst_amount': 0,
                'sgst_amount': 0,
                'rate': rate
            }
        
        invoice_data[key]['invoice_value'] += (order.invoice_amount or 0)
        invoice_data[key]['taxable_value'] += (order.taxable_value or 0)
        invoice_data[key]['igst_amount'] += (order.igst_amount or 0)
        invoice_data[key]['cgst_amount'] += (order.cgst_amount or 0)
        invoice_data[key]['sgst_amount'] += (order.sgst_amount or 0)
    
    # Subtract returns (rate-wise, use abs() since returns may store negative amounts)
    for ret in b2b_returns:
        invoice_no = ret.invoice_number or "UNKNOWN"

        # Calculate GST rate from return amounts (use abs for reliable rate calculation)
        taxable = abs(float(ret.taxable_value or 0))
        if taxable > 0:
            total_tax = abs(float(ret.igst_amount or 0)) + abs(float(ret.cgst_amount or 0)) + abs(float(ret.sgst_amount or 0))
            rate = round((total_tax / taxable * 100), 2)
        else:
            rate = 0

        key = (invoice_no, rate)
        if key in invoice_data:
            invoice_data[key]['taxable_value'] -= taxable
            invoice_data[key]['igst_amount'] -= abs(float(ret.igst_amount or 0))
            invoice_data[key]['cgst_amount'] -= abs(float(ret.cgst_amount or 0))
            invoice_data[key]['sgst_amount'] -= abs(float(ret.sgst_amount or 0))
    
    # Output B2B CSV (GSTR-1 Table 4A format with rate-wise breakdown)
    if not file_path:
        file_path = os.path.join(output_folder or "", "b2b.csv")
    
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "GSTIN of Supplier", "Trade/Legal name of the Recipient", "GSTIN/UIN of Recipient",
            "Invoice Number", "Invoice date", "Invoice Value", "Place Of Supply", "Reverse Charge",
            "Invoice Type", "E-Commerce GSTIN", "Rate", "Taxable Value", "Cess Amount"
        ])
        writer.writeheader()
        
        for (invoice_no, rate), data in sorted(invoice_data.items()):
            invoice_date_str = data['invoice_date'].strftime("%d-%m-%Y") if data['invoice_date'] else ""
            writer.writerow({
                "GSTIN of Supplier": supplier_gstin,
                "Trade/Legal name of the Recipient": data['receiver_name'],
                "GSTIN/UIN of Recipient": data['receiver_gstin'],
                "Invoice Number": invoice_no,
                "Invoice date": invoice_date_str,
                "Invoice Value": round(data['invoice_value'], 2),
                "Place Of Supply": data['place_of_supply'],
                "Reverse Charge": "N",
                "Invoice Type": "Regular",
                "E-Commerce GSTIN": "",
                "Rate": data['rate'],
                "Taxable Value": round(data['taxable_value'], 2),
                "Cess Amount": ""
            })
    
    return f"✅ B2B CSV written to {file_path} with {len(invoice_data)} invoices (Amazon B2B transactions)."


def generate_hsn_b2b_csv(financial_year, month_number, gstin_or_supplier_id, db,
                         file_path=None, output_folder=None):
    """
    Generate HSN-wise summary CSV for B2B transactions.
    
    Args:
        gstin_or_supplier_id: Either GSTIN string or legacy supplier_id integer
        
    Format: HSN, Description, UQC, Total Quantity, Total Value, Taxable Value,
            Integrated Tax Amount, Central Tax Amount, State/UT Tax Amount, Cess Amount
    """
    from models import AmazonOrder, AmazonReturn
    from datetime import datetime

    # Calculate month start and end dates
    if month_number <= 3:  # Jan-Mar
        year = financial_year
    else:  # Apr-Dec
        year = financial_year - 1

    month_start = datetime(year, month_number, 1)
    if month_number == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month_number + 1, 1)

    # Resolve supplier GSTIN for filtering
    if isinstance(gstin_or_supplier_id, str) and len(gstin_or_supplier_id) == 15:
        supplier_gstin = gstin_or_supplier_id
    else:
        supplier_gstin = get_gstin_for_supplier(gstin_or_supplier_id, db)
        if not supplier_gstin:
            raise ValueError(f"GSTIN not found for supplier ID {gstin_or_supplier_id}")

    # Query B2B transactions (where customer GSTIN is present) - filtered by seller GSTIN
    b2b_orders = db.query(AmazonOrder).filter(
        AmazonOrder.transaction_type == TransactionType.SHIPMENT,
        AmazonOrder.customer_bill_to_gstid.isnot(None),
        AmazonOrder.customer_bill_to_gstid != '',
        AmazonOrder.customer_bill_to_gstid != 'nan',
        AmazonOrder.order_date >= month_start,
        AmazonOrder.order_date < month_end,
        AmazonOrder.seller_gstin == supplier_gstin
    ).all()

    b2b_returns = db.query(AmazonReturn).filter(
        AmazonReturn.customer_bill_to_gstid.isnot(None),
        AmazonReturn.customer_bill_to_gstid != '',
        AmazonReturn.customer_bill_to_gstid != 'nan',
        AmazonReturn.order_date >= month_start,
        AmazonReturn.order_date < month_end,
        AmazonReturn.seller_gstin == supplier_gstin
    ).all()

    # Aggregate by HSN
    hsn_data = {}
    
    for order in b2b_orders:
        hsn = str(order.hsn_sac or "UNKNOWN")
        if hsn not in hsn_data:
            hsn_data[hsn] = {
                'description': order.item_description or "",
                'quantity': 0,
                'total_value': 0,
                'taxable_value': 0,
                'igst_amount': 0,
                'cgst_amount': 0,
                'sgst_amount': 0
            }
        
        hsn_data[hsn]['quantity'] += (order.quantity or 0)
        hsn_data[hsn]['total_value'] += (order.invoice_amount or 0)
        hsn_data[hsn]['taxable_value'] += (order.taxable_value or 0)
        hsn_data[hsn]['igst_amount'] += (order.igst_amount or 0)
        hsn_data[hsn]['cgst_amount'] += (order.cgst_amount or 0)
        hsn_data[hsn]['sgst_amount'] += (order.sgst_amount or 0)
    
    # Subtract returns (use abs() since returns may store negative amounts)
    for ret in b2b_returns:
        hsn = str(ret.hsn_sac or "UNKNOWN")
        if hsn in hsn_data:
            hsn_data[hsn]['quantity'] -= abs(int(ret.quantity or 0))
            hsn_data[hsn]['taxable_value'] -= abs(float(ret.taxable_value or 0))
            hsn_data[hsn]['igst_amount'] -= abs(float(ret.igst_amount or 0))
            hsn_data[hsn]['cgst_amount'] -= abs(float(ret.cgst_amount or 0))
            hsn_data[hsn]['sgst_amount'] -= abs(float(ret.sgst_amount or 0))
    
    # Output HSN B2B CSV
    if not file_path:
        file_path = os.path.join(output_folder or "", "hsn(b2b).csv")
    
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "HSN", "Description", "UQC", "Total Quantity", "Total Value",
            "Taxable Value", "Integrated Tax Amount", "Central Tax Amount",
            "State/UT Tax Amount", "Cess Amount"
        ])
        writer.writeheader()
        
        for hsn, data in sorted(hsn_data.items()):
            writer.writerow({
                "HSN": hsn,
                "Description": data['description'][:30] if data['description'] else "",  # Truncate for readability
                "UQC": "NOS",  # Unit: Numbers
                "Total Quantity": data['quantity'],
                "Total Value": round(data['total_value'], 2),
                "Taxable Value": round(data['taxable_value'], 2),
                "Integrated Tax Amount": round(data['igst_amount'], 2),
                "Central Tax Amount": round(data['cgst_amount'], 2),
                "State/UT Tax Amount": round(data['sgst_amount'], 2),
                "Cess Amount": ""
            })
    
    return f"✅ HSN (B2B) CSV written to {file_path} with {len(hsn_data)} HSN codes (Amazon B2B transactions)."


def generate_b2cl_csv(financial_year, month_number, gstin_or_supplier_id, db,
                      file_path=None, output_folder=None):
    """
    Generate B2CL (B2C Large) CSV for B2C transactions with invoice value > Rs 2.5 Lakhs.
    Complies with GSTR-1 Table 5 format.
    
    Args:
        gstin_or_supplier_id: Either GSTIN string or legacy supplier_id integer
    
    GSTR-1 Compliance:
    - Invoice-level details for B2C supplies where invoice value > 2,50,000
    - Inter-state supplies only (as per GSTR-1 requirement)
    - Rate-wise breakdown per invoice
    
    Format: Invoice Number, Invoice date, Invoice Value, Place Of Supply, Applicable % of Tax Rate,
            Rate, Taxable Value, Cess Amount, E-Commerce GSTIN
    """
    from models import AmazonOrder, AmazonReturn
    from datetime import datetime
    
    # Calculate month start and end dates
    if month_number <= 3:  # Jan-Mar
        year = financial_year
    else:  # Apr-Dec
        year = financial_year - 1
    
    month_start = datetime(year, month_number, 1)
    if month_number == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month_number + 1, 1)
    
    # Get supplier state for determining inter/intra state - accept either GSTIN string or legacy supplier_id
    if isinstance(gstin_or_supplier_id, str) and len(gstin_or_supplier_id) == 15:
        supplier_gstin = gstin_or_supplier_id
    else:
        supplier_gstin = get_gstin_for_supplier(gstin_or_supplier_id, db)
    supplier_state_code = supplier_gstin[:2] if supplier_gstin and len(supplier_gstin) >= 2 else None
    
    # Query B2C Large transactions (no customer GSTIN and invoice value > threshold)
    # Filtered by seller GSTIN to prevent cross-seller data leakage
    large_orders = db.query(AmazonOrder).filter(
        AmazonOrder.transaction_type == TransactionType.SHIPMENT,
        AmazonOrder.order_date >= month_start,
        AmazonOrder.order_date < month_end,
        AmazonOrder.seller_gstin == supplier_gstin,
        # B2C: No customer GSTIN or empty
        (AmazonOrder.customer_bill_to_gstid.is_(None) |
         (AmazonOrder.customer_bill_to_gstid == '') |
         (AmazonOrder.customer_bill_to_gstid == 'nan')),
        # Large: Invoice value > B2CL threshold
        AmazonOrder.invoice_amount > B2CL_INVOICE_THRESHOLD
    ).all()
    
    # Group by (invoice_no, rate) for rate-wise breakdown
    invoice_data = {}
    
    for order in large_orders:
        invoice_no = order.invoice_number or "UNKNOWN"
        
        # Calculate GST rate
        if order.igst_rate and order.igst_rate > 0:
            rate = round(order.igst_rate, 2)
        elif order.cgst_rate and order.sgst_rate:
            rate = round(order.cgst_rate + order.sgst_rate, 2)
        else:
            rate = 0
        
        # Determine if inter-state (GSTR-1 requirement: B2CL is for inter-state only)
        ship_to_state_code = None
        if order.ship_to_state:
            # Use helper function to get state code
            state_code_formatted = get_state_code(order.ship_to_state)
            if state_code_formatted and '-' in state_code_formatted:
                ship_to_state_code = state_code_formatted.split('-')[0]
        
        # Only include inter-state supplies (ship_to_state != supplier_state)
        if supplier_state_code and ship_to_state_code and supplier_state_code == ship_to_state_code:
            continue  # Skip intra-state supplies
        
        key = (invoice_no, rate)
        
        if key not in invoice_data:
            invoice_data[key] = {
                'invoice_date': order.invoice_date,
                'place_of_supply': get_state_code(order.ship_to_state) if order.ship_to_state else "",
                'invoice_value': 0,
                'taxable_value': 0,
                'rate': rate
            }
        
        invoice_data[key]['invoice_value'] += (order.invoice_amount or 0)
        invoice_data[key]['taxable_value'] += (order.taxable_value or 0)
    
    # Output B2CL CSV
    if not file_path:
        file_path = os.path.join(output_folder or "", "b2cl.csv")
    
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Invoice Number", "Invoice date", "Invoice Value", "Place Of Supply",
            "Applicable % of Tax Rate", "Rate", "Taxable Value", "Cess Amount", "E-Commerce GSTIN"
        ])
        writer.writeheader()
        
        for (invoice_no, rate), data in sorted(invoice_data.items()):
            invoice_date_str = data['invoice_date'].strftime("%d-%b-%y") if data['invoice_date'] else ""
            writer.writerow({
                "Invoice Number": invoice_no,
                "Invoice date": invoice_date_str,
                "Invoice Value": round(data['invoice_value'], 2),
                "Place Of Supply": data['place_of_supply'],
                "Applicable % of Tax Rate": "",  # Optional field
                "Rate": rate,
                "Taxable Value": round(data['taxable_value'], 2),
                "Cess Amount": "",
                "E-Commerce GSTIN": ""  # Blank if not e-commerce operator
            })
    
    return f"✅ B2CL CSV written to {file_path} with {len(invoice_data)} large B2C invoices (>2.5L, inter-state)."


def generate_cdnr_csv(financial_year, month_number, gstin_or_supplier_id, db,
                      file_path=None, output_folder=None):
    """
    Generate CDNR (Credit/Debit Notes - Registered) CSV for returns/adjustments to B2B customers.
    Complies with GSTR-1 Table 9B format.
    
    GSTR-1 Compliance:
    - Credit notes issued to registered persons (with GSTIN)
    - Debit notes for additional charges
    - Note-level details with reference to original supply
    
    Format: GSTIN/UIN of Recipient, Receiver Name, Note Number, Note Date, Note Type, 
            Place Of Supply, Reverse Charge, Note Supply Type, Note Value, Applicable % of Tax Rate,
            Rate, Taxable Value, Cess Amount
    """
    from models import AmazonReturn
    from datetime import datetime

    # Calculate month start and end dates
    if month_number <= 3:  # Jan-Mar
        year = financial_year
    else:  # Apr-Dec
        year = financial_year - 1

    month_start = datetime(year, month_number, 1)
    if month_number == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month_number + 1, 1)

    # Resolve supplier GSTIN for filtering
    if isinstance(gstin_or_supplier_id, str) and len(gstin_or_supplier_id) == 15:
        supplier_gstin = gstin_or_supplier_id
    else:
        supplier_gstin = get_gstin_for_supplier(gstin_or_supplier_id, db)
        if not supplier_gstin:
            raise ValueError(f"GSTIN not found for supplier ID {gstin_or_supplier_id}")

    # Query returns to B2B customers (those with GSTIN) - filtered by seller GSTIN
    b2b_returns = db.query(AmazonReturn).filter(
        AmazonReturn.transaction_type.in_([TransactionType.REFUND, TransactionType.CANCEL]),
        AmazonReturn.customer_bill_to_gstid.isnot(None),
        AmazonReturn.customer_bill_to_gstid != '',
        AmazonReturn.customer_bill_to_gstid != 'nan',
        AmazonReturn.order_date >= month_start,
        AmazonReturn.order_date < month_end,
        AmazonReturn.seller_gstin == supplier_gstin
    ).all()
    
    # Group by (invoice_no, rate) for rate-wise breakdown
    note_data = {}
    
    for ret in b2b_returns:
        # Generate note number using helper function
        note_no = generate_note_number(
            ret.invoice_number or ret.order_id,
            NoteType.CREDIT
        )
        
        # Calculate GST rate from return amounts
        taxable = abs(ret.taxable_value or 0)
        if taxable > 0:
            total_tax = abs(ret.igst_amount or 0) + abs(ret.cgst_amount or 0) + abs(ret.sgst_amount or 0)
            rate = round((total_tax / taxable * 100), 2)
        else:
            rate = 0
        
        key = (note_no, rate)
        
        if key not in note_data:
            note_data[key] = {
                'receiver_gstin': ret.customer_bill_to_gstid or "",
                'receiver_name': ret.buyer_name or "",
                'note_date': ret.invoice_date,
                'note_type': NoteType.CREDIT,  # Credit Note for returns
                'place_of_supply': get_state_code(ret.ship_to_state) if ret.ship_to_state else "",
                'note_value': 0,
                'taxable_value': 0,
                'rate': rate
            }
        
        # Return amounts are typically negative, so take absolute value
        note_data[key]['note_value'] += abs(ret.return_amount or 0)
        note_data[key]['taxable_value'] += abs(ret.taxable_value or 0)
    
    # Output CDNR CSV
    if not file_path:
        file_path = os.path.join(output_folder or "", "cdnr.csv")
    
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "GSTIN/UIN of Recipient", "Receiver Name", "Note Number", "Note Date", "Note Type",
            "Place Of Supply", "Reverse Charge", "Note Supply Type", "Note Value",
            "Applicable % of Tax Rate", "Rate", "Taxable Value", "Cess Amount"
        ])
        writer.writeheader()
        
        for (note_no, rate), data in sorted(note_data.items()):
            note_date_str = data['note_date'].strftime("%d-%b-%y") if data['note_date'] else ""
            writer.writerow({
                "GSTIN/UIN of Recipient": data['receiver_gstin'],
                "Receiver Name": data['receiver_name'],
                "Note Number": note_no,
                "Note Date": note_date_str,
                "Note Type": data['note_type'],  # C or D
                "Place Of Supply": data['place_of_supply'],
                "Reverse Charge": "N",
                "Note Supply Type": "Regular B2B",
                "Note Value": round(data['note_value'], 2),
                "Applicable % of Tax Rate": "",
                "Rate": rate,
                "Taxable Value": round(data['taxable_value'], 2),
                "Cess Amount": ""
            })
    
    return f"✅ CDNR CSV written to {file_path} with {len(note_data)} credit/debit notes (B2B returns)."


def generate_gstr1_excel_workbook(financial_year, month_number, gstin_or_supplier_id, db,
                                   file_path=None, output_folder=None):
    """
    Generate comprehensive GSTR-1 Excel Workbook with all tables in separate sheets.
    This creates a single Excel file similar to the official GSTR1_Excel_Workbook_Template.
    
    Args:
        gstin_or_supplier_id: Either GSTIN string or legacy supplier_id integer
    
    Includes all major GSTR-1 tables:
    - B2B (Table 4A, 4B, 4C) - Business to Business supplies
    - B2CL (Table 5) - Large B2C invoices (>2.5L, inter-state)
    - B2CS (Table 7) - Small B2C invoices (consolidated)
    - CDNR (Table 9B) - Credit/Debit notes for registered persons
    - HSN (Table 12) - HSN-wise summary
    - DOCS (Table 13) - Document details
    
    Args:
        financial_year: FY starting year (e.g., 2024 for FY 2024-25)
        month_number: Month (1-12)
        supplier_id: Supplier database ID
        db: SQLAlchemy database session
        file_path: Optional custom output file path
        output_folder: Optional output directory
        
    Returns:
        Success message with file path and summary
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import tempfile
    import csv as csv_module
    
    # Determine output file path
    if not file_path:
        # Calculate calendar year from financial year and month
        # FY 2025 = Apr 2025 to Mar 2026
        # Months 4-12: Use financial_year (e.g., Oct 2025 for FY 2025)
        # Months 1-3: Use financial_year + 1 (e.g., Jan 2026 for FY 2025)
        if month_number <= 3:
            year = financial_year + 1
        else:
            year = financial_year
        
        month_name = datetime(year, month_number, 1).strftime("%B")
        file_name = f"GSTR1_FY{financial_year}_{month_name}_{year}.xlsx"
        file_path = os.path.join(output_folder or "", file_name)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create a summary sheet as the default/first sheet to avoid the openpyxl warning
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = f"GSTR-1 Report"
    ws_summary['A2'] = f"Financial Year: {financial_year}"
    ws_summary['A3'] = f"Month: {month_number}"
    
    # Define styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header_row(ws, row_num=1):
        """Apply header styling to a row."""
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def auto_adjust_column_width(ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generate each CSV in memory and add to workbook
    table_count = 0
    total_records = 0
    
    # 1. B2B Sheet
    try:
        temp_csv = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='', encoding='utf-8')
        temp_csv_path = temp_csv.name
        temp_csv.close()
        
        result = generate_b2b_csv(financial_year, month_number, gstin_or_supplier_id, db, file_path=temp_csv_path)
        
        # Read CSV and add to Excel
        ws = wb.create_sheet("B2B")
        with open(temp_csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv_module.reader(csvfile)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.border = thin_border
        
        style_header_row(ws, 1)
        auto_adjust_column_width(ws)
        os.unlink(temp_csv_path)
        
        # Extract record count from result message
        import re
        match = re.search(r'with (\d+)', result)
        if match:
            count = int(match.group(1))
            total_records += count
        table_count += 1
    except Exception as e:
        print(f"⚠️  Warning: Could not generate B2B sheet: {e}")
    
    # 2. B2CL Sheet
    try:
        temp_csv = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='', encoding='utf-8')
        temp_csv_path = temp_csv.name
        temp_csv.close()
        
        result = generate_b2cl_csv(financial_year, month_number, gstin_or_supplier_id, db, file_path=temp_csv_path)
        
        ws = wb.create_sheet("B2CL")
        with open(temp_csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv_module.reader(csvfile)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        style_header_row(ws, 1)
        auto_adjust_column_width(ws)
        os.unlink(temp_csv_path)
        
        match = re.search(r'with (\d+)', result)
        if match:
            count = int(match.group(1))
            total_records += count
        table_count += 1
    except Exception as e:
        print(f"⚠️  Warning: Could not generate B2CL sheet: {e}")
    
    # 3. B2CS Sheet (B2C Small - using generate_gst_pivot_csv)
    try:
        temp_csv = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='', encoding='utf-8')
        temp_csv_path = temp_csv.name
        temp_csv.close()
        
        # B2CS data is generated by generate_gst_pivot_csv function
        result = generate_gst_pivot_csv(financial_year, month_number, gstin_or_supplier_id, db, file_path=temp_csv_path)
        
        ws = wb.create_sheet("B2CS")
        with open(temp_csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv_module.reader(csvfile)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        style_header_row(ws, 1)
        auto_adjust_column_width(ws)
        os.unlink(temp_csv_path)
        
        match = re.search(r'with (\d+)', result)
        if match:
            count = int(match.group(1))
            total_records += count
        table_count += 1
    except Exception as e:
        print(f"⚠️  Warning: Could not generate B2CS sheet: {e}")
    
    # 4. CDNR Sheet
    try:
        temp_csv = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='', encoding='utf-8')
        temp_csv_path = temp_csv.name
        temp_csv.close()
        
        result = generate_cdnr_csv(financial_year, month_number, gstin_or_supplier_id, db, file_path=temp_csv_path)
        
        ws = wb.create_sheet("CDNR")
        with open(temp_csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv_module.reader(csvfile)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        style_header_row(ws, 1)
        auto_adjust_column_width(ws)
        os.unlink(temp_csv_path)
        
        match = re.search(r'with (\d+)', result)
        if match:
            count = int(match.group(1))
            total_records += count
        table_count += 1
    except Exception as e:
        print(f"⚠️  Warning: Could not generate CDNR sheet: {e}")
    
    # 5. HSN Summary Sheet
    try:
        temp_csv = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='', encoding='utf-8')
        temp_csv_path = temp_csv.name
        temp_csv.close()
        
        result = generate_hsn_b2b_csv(financial_year, month_number, gstin_or_supplier_id, db, file_path=temp_csv_path)
        
        ws = wb.create_sheet("HSN")
        with open(temp_csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv_module.reader(csvfile)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        style_header_row(ws, 1)
        auto_adjust_column_width(ws)
        os.unlink(temp_csv_path)
        
        match = re.search(r'with (\d+)', result)
        if match:
            count = int(match.group(1))
            total_records += count
        table_count += 1
    except Exception as e:
        print(f"⚠️  Warning: Could not generate HSN sheet: {e}")
    
    # 6. Add summary/index sheet as first sheet
    ws_summary = wb.create_sheet("Summary", 0)
    
    # Add summary information
    # Financial year runs April to March
    # April-December (months 4-12): Use financial_year as calendar year
    # January-March (months 1-3): Use financial_year + 1 as calendar year
    if month_number <= 3:
        year = financial_year + 1  # Jan-Mar are in next calendar year
    else:
        year = financial_year  # Apr-Dec are in same calendar year
    
    month_name = datetime(year, month_number, 1).strftime("%B %Y")
    
    summary_data = [
        ["GSTR-1 Return Summary"],
        [""],
        ["Financial Year:", f"FY {financial_year}-{financial_year+1}"],
        ["Period:", month_name],
        ["Generated On:", datetime.now().strftime("%d-%b-%Y %H:%M:%S")],
        [""],
        ["Tables Included:", ""],
        ["  - B2B (Business to Business)", "Invoices to registered persons"],
        ["  - B2CL (B2C Large)", "B2C invoices >₹2.5L (inter-state)"],
        ["  - B2CS (B2C Small)", "Consolidated B2C supplies"],
        ["  - CDNR (Credit/Debit Notes)", "Notes for registered persons"],
        ["  - HSN Summary", "HSN-wise summary"],
        [""],
        ["Statistics:", ""],
        [f"  Total Tables Generated:", table_count],
        [f"  Total Records:", total_records],
        [""],
        ["Instructions:", ""],
        ["  1. Review each sheet for accuracy"],
        ["  2. Upload to GST portal offline tool"],
        ["  3. Validate before filing"],
        [""],
        ["⚠️ Note:", "This is auto-generated data. Please verify before filing."]
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            
            # Style first row as title
            if row_idx == 1:
                cell.font = Font(bold=True, size=16, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Style section headers
            elif col_idx == 1 and value and value.endswith(":") and len(value) > 5:
                cell.font = Font(bold=True, size=11)
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 50
    
    # Save workbook
    wb.save(file_path)
    
    return f"✅ GSTR-1 Excel Workbook created: {file_path}\n   📊 {table_count} tables with {total_records} total records"
